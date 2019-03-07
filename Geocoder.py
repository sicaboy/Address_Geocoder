
print ('Developed by Ashish Baboo.' + '\n')

import xlrd
import geocoder
import openpyxl
from openpyxl import load_workbook




location = "address_to_be_processed.xlsx"
workbook = xlrd.open_workbook(location)
sheet = workbook.sheet_by_index(1)


Rows = sheet.nrows
Cols = sheet.ncols

wb = load_workbook('address_to_be_processed.xlsx')
ws1 = wb.get_sheet_by_name("Locations")


def ascii_rem(text):

    return ''.join(i for i in text if ord(i)<128)


for i in range(1,Rows):

    add = ""

    if ws1.cell(row = i+1, column = 4).value is not None:
        add = ''.join(str([ws1.cell(row = i+1, column = 4).value]))
    if ws1.cell(row = i+1, column = 5).value is not None:
        add = ''.join([add,",",str(ws1.cell(row = i+1, column = 5).value)]) 
    if ws1.cell(row = i+1, column = 6).value is not None:
        add = ''.join([add,",",str(ws1.cell(row = i+1, column = 6).value)])
    if ws1.cell(row = i+1, column = 7).value is not None:
        add = ''.join([add,",",str(ws1.cell(row = i+1, column = 7).value)])
    if ws1.cell(row = i+1, column = 8).value is not None:
        add = ''.join([add,",",str(ws1.cell(row = i+1, column = 8).value)])
    if ws1.cell(row = i+1, column = 9).value is not None:
        add = ''.join([add,",",str(ws1.cell(row = i+1, column = 9).value)])

    add_cleaned = ascii_rem(add)

    cadd = geocoder.google(add_cleaned  , language = "en" , key = "your key here")
    print(i , cadd.latlng , cadd.quality ,cadd.country)

    if cadd.country is None:
        add = ""
        
        if ws1.cell(row = i+1, column = 5).value is not None:
            add = ''.join(str([add,",",ws1.cell(row = i+1, column = 5).value]))
        if ws1.cell(row = i+1, column = 6).value is not None:
            add = ''.join([add,",",str(ws1.cell(row = i+1, column = 6).value)])
        if ws1.cell(row = i+1, column = 7).value is not None:
            add = ''.join([add,",",str(ws1.cell(row = i+1, column = 7).value)])
        if ws1.cell(row = i+1, column = 8).value is not None:
            add = ''.join([add,",",str(ws1.cell(row = i+1, column = 8).value)])
        if ws1.cell(row = i+1, column = 9).value is not None:
            add = ''.join([add,",",str(ws1.cell(row = i+1, column = 9).value)])

        add_cleaned = ascii_rem(add)

        cadd = geocoder.google(add_cleaned  , language = "en" , key = "Your key here")
        print(i , cadd.latlng , cadd.quality ,cadd.country)

        if cadd.country is None:
            add = ""
            
            if ws1.cell(row = i+1, column = 7).value is not None:
                add = ''.join([add,",",str(ws1.cell(row = i+1, column = 7).value)])
            if ws1.cell(row = i+1, column = 8).value is not None:
                add = ''.join([add,",",str(ws1.cell(row = i+1, column = 8).value)])
            if ws1.cell(row = i+1, column = 9).value is not None:
                add = ''.join([add,",",str(ws1.cell(row = i+1, column = 9).value)])

            add_cleaned = ascii_rem(add)
            

            cadd = geocoder.google(add_cleaned  , language = "en" , key = "AIzaSyDn53t2Z5bGOBeL-iK8ARWJhuT8o-mspjc")
            print(i , cadd.latlng , cadd.quality ,cadd.country)

    S_no = ws1.cell(row = i+1, column = 1) 
    O_Housenumber = ws1.cell(row = i+1, column = 20)
    O_Street = ws1.cell(row = i+1, column = 21)
    O_City = ws1.cell(row = i+1, column = 22)
    O_County = ws1.cell(row = i+1, column = 23)
    O_State = ws1.cell(row = i+1, column = 24)
    O_Postal = ws1.cell(row = i+1, column = 25)
    O_Country = ws1.cell(row = i+1, column = 26)
    Latitude = ws1.cell(row = i+1, column = 10)
    Longitude = ws1.cell(row = i+1, column = 11)
    Accuracy = ws1.cell(row = i+1, column = 12)
    S_no.value = i
    O_Housenumber.value = cadd.housenumber
    O_Street.value = cadd.street 
    O_City.value = cadd.city
    O_County.value = cadd.county
    O_State.value = cadd.state_long
    O_Postal.value = cadd.postal
    O_Country.value = cadd.country_long
    Latitude.value = cadd.lat
    Longitude.value = cadd.lng
    Accuracy.value = cadd.quality
    
    wb.save('address_to_be_processed.xlsx')
    
        



    


